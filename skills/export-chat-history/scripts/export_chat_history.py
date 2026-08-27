#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import urllib.error
import urllib.request
import uuid
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Any, NoReturn

try:
    from zoneinfo import ZoneInfo
except ImportError:  # pragma: no cover - Python 3.8 fallback
    ZoneInfo = None  # type: ignore[assignment,misc]

sys.stderr = sys.stdout

MAX_EXPORT_ROWS = 50_000
MAX_EXCEL_FILE_SIZE = 25 * 1024 * 1024
EXCEL_CELL_MAX_LENGTH = 32_767
ILLEGAL_XML_CHARACTERS = re.compile(
    r"[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F-\x84\x86-\x9F]"
)
DATETIME_FORMATS = (
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%dT%H:%M",
    "%Y-%m-%dT%H:%M:%S",
)


class SkillArgumentParser(argparse.ArgumentParser):
    def error(self, message: str) -> NoReturn:
        raise ValueError(f"参数错误：{message}")


def _shanghai_timezone():
    if ZoneInfo is not None:
        try:
            return ZoneInfo("Asia/Shanghai")
        except Exception:
            pass
    return timezone(timedelta(hours=8), name="Asia/Shanghai")


SHANGHAI_TZ = _shanghai_timezone()


def _skill_root() -> Path:
    return Path(__file__).resolve().parent.parent


def _skill_venv_python() -> Path:
    venv_dir = _skill_root() / ".venv"
    if sys.platform == "win32":
        return venv_dir / "Scripts" / "python.exe"
    return venv_dir / "bin" / "python"


def _python_executable() -> str:
    if sys.executable:
        return sys.executable
    for candidate in ("python3", "python"):
        executable = shutil.which(candidate)
        if executable:
            return executable
    raise RuntimeError("无法找到 Python 解释器路径")


def _run_bootstrap() -> None:
    bootstrap = Path(__file__).resolve().parent / "bootstrap.py"
    result = subprocess.run(
        [_python_executable(), str(bootstrap)],
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
    )
    if result.returncode != 0:
        output = result.stdout.strip()
        detail = f"：{output}" if output else ""
        raise RuntimeError(
            f"安装技能依赖失败，退出码: {result.returncode}{detail}"
        )


def _ensure_runtime_dependencies() -> None:
    try:
        import openpyxl  # noqa: F401
        import pymysql  # noqa: F401

        return
    except ModuleNotFoundError:
        pass

    _run_bootstrap()
    venv_python = _skill_venv_python()
    if not venv_python.is_file():
        raise RuntimeError("安装依赖后仍未找到技能虚拟环境")

    venv_dir = (_skill_root() / ".venv").resolve()
    if Path(sys.prefix).resolve() == venv_dir:
        try:
            import openpyxl  # noqa: F401
            import pymysql  # noqa: F401

            return
        except ModuleNotFoundError as exc:
            raise RuntimeError(
                "技能虚拟环境缺少 pymysql 或 openpyxl"
            ) from exc

    os.execv(
        str(venv_python),
        [str(venv_python), str(Path(__file__).resolve()), *sys.argv[1:]],
    )


def _parse_args(argv: list[str]) -> argparse.Namespace:
    parser = SkillArgumentParser(description="导出当前群聊的聊天记录")
    parser.add_argument("--date", dest="export_date", default="")
    parser.add_argument(
        "--start-time",
        "--start_time",
        dest="start_time",
        default="",
    )
    parser.add_argument(
        "--end-time",
        "--end_time",
        dest="end_time",
        default="",
    )
    args = parser.parse_args(argv)

    if args.export_date and (args.start_time or args.end_time):
        raise ValueError("date 不能与 start_time 或 end_time 同时使用")
    if bool(args.start_time) != bool(args.end_time):
        raise ValueError("自定义时间范围必须同时提供 start_time 和 end_time")
    return args


def _parse_datetime(value: str, field_name: str) -> datetime:
    text = value.strip()
    for pattern in DATETIME_FORMATS:
        try:
            return datetime.strptime(text, pattern).replace(tzinfo=SHANGHAI_TZ)
        except ValueError:
            continue
    raise ValueError(
        f"{field_name} 必须使用 YYYY-MM-DD HH:mm 或 "
        "YYYY-MM-DD HH:mm:ss 格式"
    )


def _resolve_time_range(
    args: argparse.Namespace,
    now: datetime | None = None,
) -> tuple[datetime, datetime]:
    current = now or datetime.now(SHANGHAI_TZ)
    if current.tzinfo is None:
        current = current.replace(tzinfo=SHANGHAI_TZ)
    else:
        current = current.astimezone(SHANGHAI_TZ)

    if args.export_date:
        try:
            selected_date = date.fromisoformat(args.export_date.strip())
        except ValueError as exc:
            raise ValueError("date 必须使用 YYYY-MM-DD 格式") from exc
        start = datetime.combine(
            selected_date,
            time.min,
            tzinfo=SHANGHAI_TZ,
        )
        end = start + timedelta(days=1)
    elif args.start_time and args.end_time:
        start = _parse_datetime(args.start_time, "start_time")
        end = _parse_datetime(args.end_time, "end_time")
    else:
        start = current.replace(hour=0, minute=0, second=0, microsecond=0)
        end = current

    if end <= start:
        raise ValueError("结束时间必须晚于开始时间")
    return start, end


def _require_env(name: str) -> str:
    value = os.environ.get(name, "").strip()
    if not value:
        raise ValueError(f"环境变量 {name} 未配置")
    return value


def _conversation_context() -> tuple[str, str, str]:
    chat_room_id = _require_env("ROBOT_FROM_WX_ID")
    if not chat_room_id.endswith("@chatroom"):
        raise ValueError("聊天记录导出仅支持在微信群聊中使用")
    self_wx_id = _require_env("ROBOT_WX_ID")
    client_port = _require_env("ROBOT_WECHAT_CLIENT_PORT")
    if not client_port.isdigit():
        raise ValueError("环境变量 ROBOT_WECHAT_CLIENT_PORT 格式不正确")
    return chat_room_id, self_wx_id, client_port


def _mysql_connect() -> Any:
    try:
        import pymysql
    except ModuleNotFoundError as exc:  # pragma: no cover - runtime guard
        raise RuntimeError("pymysql 依赖未安装") from exc

    try:
        port = int(os.environ.get("MYSQL_PORT", "3306"))
    except ValueError as exc:
        raise ValueError("环境变量 MYSQL_PORT 格式不正确") from exc

    database = _require_env("ROBOT_CODE")
    return pymysql.connect(
        host=os.environ.get("MYSQL_HOST", "127.0.0.1"),
        port=port,
        user=os.environ.get("MYSQL_USER", "root"),
        password=os.environ.get("MYSQL_PASSWORD", ""),
        database=database,
        charset="utf8mb4",
        connect_timeout=10,
        read_timeout=120,
        cursorclass=pymysql.cursors.DictCursor,
    )


def _fetch_chat_room_name(connection: Any, chat_room_id: str) -> str:
    sql = """
        SELECT nickname
        FROM contacts
        WHERE wechat_id = %s
            AND deleted_at IS NULL
        LIMIT 1
    """
    with connection.cursor() as cursor:
        cursor.execute(sql, (chat_room_id,))
        row = cursor.fetchone()
    if isinstance(row, dict):
        nickname = str(row.get("nickname") or "").strip()
        if nickname:
            return nickname
    return chat_room_id


def _fetch_messages(
    connection: Any,
    chat_room_id: str,
    self_wx_id: str,
    start_timestamp: int,
    end_timestamp: int,
) -> list[dict[str, Any]]:
    sql = """
        SELECT
            messages.id,
            messages.sender_wxid,
            COALESCE(
                NULLIF(chat_room_members.remark, ''),
                NULLIF(chat_room_members.nickname, ''),
                messages.sender_wxid
            ) AS nickname,
            CASE
                WHEN messages.type = 49 THEN
                    CASE
                        WHEN EXTRACTVALUE(
                            messages.content,
                            '/msg/appmsg/type'
                        ) = '57' THEN EXTRACTVALUE(
                            messages.content,
                            '/msg/appmsg/title'
                        )
                        WHEN EXTRACTVALUE(
                            messages.content,
                            '/msg/appmsg/type'
                        ) IN ('4', '5') THEN CONCAT(
                            '网页分享消息，标题: ',
                            EXTRACTVALUE(
                                messages.content,
                                '/msg/appmsg/title'
                            ),
                            '，描述：',
                            EXTRACTVALUE(
                                messages.content,
                                '/msg/appmsg/des'
                            )
                        )
                        WHEN EXTRACTVALUE(
                            messages.content,
                            '/msg/appmsg/type'
                        ) = '6' THEN CONCAT(
                            '文件消息，文件名: ',
                            EXTRACTVALUE(
                                messages.content,
                                '/msg/appmsg/title'
                            )
                        )
                        ELSE EXTRACTVALUE(
                            messages.content,
                            '/msg/appmsg/des'
                        )
                    END
                ELSE messages.content
            END AS message,
            messages.created_at
        FROM messages
        LEFT JOIN chat_room_members
            ON chat_room_members.wechat_id = messages.sender_wxid
            AND chat_room_members.chat_room_id = messages.from_wxid
        WHERE messages.from_wxid = %s
            AND (
                messages.type = 1
                OR (
                    messages.type = 49
                    AND EXTRACTVALUE(
                        messages.content,
                        '/msg/appmsg/type'
                    ) IN ('57', '4', '5', '6')
                )
            )
            AND messages.sender_wxid != %s
            AND messages.created_at >= %s
            AND messages.created_at < %s
        ORDER BY messages.created_at ASC, messages.id ASC
        LIMIT %s
    """
    with connection.cursor() as cursor:
        cursor.execute(
            sql,
            (
                chat_room_id,
                self_wx_id,
                start_timestamp,
                end_timestamp,
                MAX_EXPORT_ROWS + 1,
            ),
        )
        rows = cursor.fetchall()

    if len(rows) > MAX_EXPORT_ROWS:
        raise ValueError(
            f"聊天记录超过 {MAX_EXPORT_ROWS:,} 条，请缩短导出时间范围"
        )
    return [row for row in rows if isinstance(row, dict)]


def _safe_excel_text(value: Any) -> str:
    text_value = "" if value is None else str(value)
    text_value = ILLEGAL_XML_CHARACTERS.sub("", text_value)
    prefix = "'" if text_value.startswith(("=", "+", "-", "@")) else ""
    maximum_text_length = EXCEL_CELL_MAX_LENGTH - len(prefix)
    if len(text_value) > maximum_text_length:
        suffix = "\n……（内容过长，已截断）"
        text_value = text_value[
            : maximum_text_length - len(suffix)
        ] + suffix
    return prefix + text_value


def _safe_filename_part(value: str) -> str:
    sanitized = re.sub(r'[\\/:*?"<>|\x00-\x1F]', "_", value)
    sanitized = re.sub(r"\s+", "_", sanitized).strip("._")
    return sanitized[:40] or "群聊"


def _create_output_path(
    chat_room_name: str,
    start: datetime,
    end: datetime,
) -> Path:
    output_dir = Path(tempfile.gettempdir()) / "wechat-robot-chat-history"
    output_dir.mkdir(mode=0o700, parents=True, exist_ok=True)
    try:
        output_dir.chmod(0o700)
    except OSError:
        pass

    filename = (
        f"聊天记录_{_safe_filename_part(chat_room_name)}_"
        f"{start:%Y%m%d_%H%M}-{end:%Y%m%d_%H%M}_"
        f"{uuid.uuid4().hex[:8]}.xlsx"
    )
    return output_dir / filename


def _build_workbook(
    output_path: Path,
    chat_room_id: str,
    chat_room_name: str,
    start: datetime,
    end: datetime,
    messages: list[dict[str, Any]],
    exported_at: datetime | None = None,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.cell import WriteOnlyCell
        from openpyxl.styles import Alignment, Font, PatternFill
    except ModuleNotFoundError as exc:  # pragma: no cover - runtime guard
        raise RuntimeError("openpyxl 依赖未安装") from exc

    workbook = Workbook(write_only=True)
    workbook.properties.creator = "wechat-robot"
    workbook.properties.title = f"{chat_room_name}聊天记录"
    workbook.properties.subject = "微信群聊记录导出"

    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="5B9BD5")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    header_font = Font(color="FFFFFF", bold=True)
    label_font = Font(bold=True, color="1F4E78")
    wrapped_alignment = Alignment(
        vertical="top",
        wrap_text=True,
    )
    centered_alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    info_sheet = workbook.create_sheet("导出说明")
    info_sheet.column_dimensions["A"].width = 18
    info_sheet.column_dimensions["B"].width = 72

    info_title = WriteOnlyCell(info_sheet, value="群聊记录导出说明")
    info_title.fill = title_fill
    info_title.font = title_font
    info_title.alignment = centered_alignment
    info_sheet.append([info_title, ""])

    current = exported_at or datetime.now(SHANGHAI_TZ)
    info_rows = (
        ("群聊名称", chat_room_name),
        ("群聊 ID", chat_room_id),
        ("开始时间（包含）", start.strftime("%Y-%m-%d %H:%M:%S")),
        ("结束时间（不包含）", end.strftime("%Y-%m-%d %H:%M:%S")),
        ("消息数量", len(messages)),
        ("导出时间", current.strftime("%Y-%m-%d %H:%M:%S")),
    )
    for label, value in info_rows:
        label_cell = WriteOnlyCell(info_sheet, value=label)
        label_cell.font = label_font
        value_cell = WriteOnlyCell(
            info_sheet,
            value=_safe_excel_text(value),
        )
        value_cell.alignment = wrapped_alignment
        info_sheet.append([label_cell, value_cell])

    history_sheet = workbook.create_sheet("聊天记录")
    history_sheet.freeze_panes = "A2"
    widths = {"A": 9, "B": 21, "C": 24, "D": 30, "E": 88}
    for column, width in widths.items():
        history_sheet.column_dimensions[column].width = width

    headers = (
        "序号",
        "发送时间",
        "发送人",
        "发送人微信 ID",
        "消息内容",
    )
    header_cells = []
    for header in headers:
        cell = WriteOnlyCell(history_sheet, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = centered_alignment
        header_cells.append(cell)
    history_sheet.append(header_cells)

    for index, message in enumerate(messages, start=1):
        created_at = int(message.get("created_at") or 0)
        sent_at = datetime.fromtimestamp(
            created_at,
            tz=SHANGHAI_TZ,
        ).strftime("%Y-%m-%d %H:%M:%S")
        row_values = (
            index,
            sent_at,
            _safe_excel_text(message.get("nickname")),
            _safe_excel_text(message.get("sender_wxid")),
            _safe_excel_text(message.get("message")),
        )
        row_cells = []
        for column_index, value in enumerate(row_values):
            cell = WriteOnlyCell(history_sheet, value=value)
            cell.alignment = (
                centered_alignment
                if column_index in (0, 1)
                else wrapped_alignment
            )
            row_cells.append(cell)
        history_sheet.append(row_cells)

    history_sheet.auto_filter.ref = f"A1:E{len(messages) + 1}"

    temporary_path = output_path.with_name(
        f".{output_path.stem}.{uuid.uuid4().hex}.tmp.xlsx"
    )
    try:
        workbook.save(temporary_path)
        if temporary_path.stat().st_size > MAX_EXCEL_FILE_SIZE:
            raise ValueError("生成的 Excel 文件超过 25MB，请缩短导出时间范围")
        temporary_path.replace(output_path)
    finally:
        temporary_path.unlink(missing_ok=True)


def _raise_for_client_error(payload: dict[str, Any]) -> None:
    code = payload.get("code")
    if code is not None and str(code) != "200":
        message = str(payload.get("message") or f"客户端返回错误码 {code}")
        raise RuntimeError(message)


def _client_private_token() -> str:
    return os.environ.get("ROBOT_CLIENT_PRIVATE_TOKEN", "").strip()


def _send_file(client_port: str, chat_room_id: str, path: Path) -> None:
    url = (
        f"http://127.0.0.1:{client_port}"
        "/api/v1/robot/message/send/file/local"
    )
    body = json.dumps(
        {"to_wxid": chat_room_id, "file_path": str(path)}
    ).encode("utf-8")
    request = urllib.request.Request(
        url,
        data=body,
        headers={
            "Content-Type": "application/json",
            "X-Private-Token": _client_private_token(),
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=300) as response:
            response_text = response.read().decode("utf-8")
    except urllib.error.HTTPError as exc:
        error_body = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"发送 Excel 失败，HTTP {exc.code}: {error_body}") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"发送 Excel 失败: {exc.reason}") from exc

    if not response_text.strip():
        return
    try:
        payload = json.loads(response_text)
    except json.JSONDecodeError as exc:
        raise RuntimeError("客户端文件发送接口返回了无效 JSON") from exc
    if not isinstance(payload, dict):
        raise RuntimeError("客户端文件发送接口返回格式不正确")
    _raise_for_client_error(payload)


def _emit_success(
    chat_room_name: str,
    start: datetime,
    end: datetime,
    message_count: int,
) -> None:
    result = {
        "ok": True,
        "message": "聊天记录已导出为 Excel 并发送到当前群聊",
        "chat_room_name": chat_room_name,
        "start_time": start.strftime("%Y-%m-%d %H:%M:%S"),
        "end_time": end.strftime("%Y-%m-%d %H:%M:%S"),
        "message_count": message_count,
    }
    sys.stdout.write(json.dumps(result, ensure_ascii=False) + "\n")


def main() -> int:
    output_path: Path | None = None
    file_sent = False
    try:
        args = _parse_args(sys.argv[1:])
        start, end = _resolve_time_range(args)
        chat_room_id, self_wx_id, client_port = _conversation_context()
        _ensure_runtime_dependencies()

        connection = _mysql_connect()
        try:
            chat_room_name = _fetch_chat_room_name(
                connection,
                chat_room_id,
            )
            messages = _fetch_messages(
                connection,
                chat_room_id,
                self_wx_id,
                int(start.timestamp()),
                int(end.timestamp()),
            )
        finally:
            connection.close()

        if not messages:
            raise ValueError("该时间范围没有可导出的聊天记录")

        output_path = _create_output_path(chat_room_name, start, end)
        _build_workbook(
            output_path,
            chat_room_id,
            chat_room_name,
            start,
            end,
            messages,
        )
        _send_file(client_port, chat_room_id, output_path)
        file_sent = True
        _emit_success(
            chat_room_name,
            start,
            end,
            len(messages),
        )
        return 0
    except (ValueError, RuntimeError) as exc:
        sys.stdout.write(f"导出聊天记录失败: {exc}\n")
        return 1
    except Exception as exc:
        sys.stdout.write(f"导出聊天记录失败: {exc}\n")
        return 1
    finally:
        if file_sent and output_path is not None:
            output_path.unlink(missing_ok=True)


if __name__ == "__main__":
    raise SystemExit(main())
