#!/usr/bin/env python3

from __future__ import annotations

import argparse
import os
import re
import tempfile
from copy import copy
from pathlib import Path
from typing import Any, Iterable, Optional

from _xlsx_common import (
    EXCEL_INPUT_SUFFIXES,
    EXCEL_OUTPUT_SUFFIXES,
    SkillArgumentParser,
    input_file,
    load_json_argument,
    normalize_formula_error,
    openpyxl_load_options,
    output_file,
    publish_file,
    run_cli,
    validate_cell_range,
    validate_cell_reference,
    workbook_has_external_links,
)


MAX_OPERATIONS = 500
MAX_WRITTEN_CELLS = 100_000
HEX_COLOR_RE = re.compile(r"^(?:[0-9A-Fa-f]{6}|[0-9A-Fa-f]{8})$")
TABLE_NAME_RE = re.compile(r"^[A-Za-z_\\][A-Za-z0-9_.\\]*$")
RISKY_FORMULA_RE = re.compile(
    r"\b(?:XLOOKUP|XMATCH|SORT|FILTER|UNIQUE|SEQUENCE)\s*\(",
    re.IGNORECASE,
)


def _expect_object(value: Any, label: str) -> dict[str, Any]:
    if not isinstance(value, dict):
        raise ValueError(f"{label} 必须是对象")
    return value


def _expect_list(value: Any, label: str) -> list[Any]:
    if not isinstance(value, list):
        raise ValueError(f"{label} 必须是数组")
    return value


def _color(value: Any, label: str) -> str:
    if not isinstance(value, str) or not HEX_COLOR_RE.fullmatch(value):
        raise ValueError(f"{label} 必须是 6 位或 8 位十六进制颜色")
    return value.upper()


def _sheet(workbook: Any, name: Any) -> Any:
    if not isinstance(name, str) or not name:
        raise ValueError("操作必须提供非空 sheet")
    if name not in workbook.sheetnames:
        raise ValueError(f"工作表不存在：{name}")
    return workbook[name]


def _iter_range_cells(worksheet: Any, reference: str) -> Iterable[Any]:
    normalized = validate_cell_range(reference)
    if ":" not in normalized:
        yield worksheet[normalized]
        return
    for row in worksheet[normalized]:
        for cell in row:
            yield cell


def _apply_font(cell: Any, spec: dict[str, Any]) -> None:
    font = copy(cell.font)
    allowed = {
        "name",
        "size",
        "bold",
        "italic",
        "underline",
        "strike",
        "color",
        "vert_align",
    }
    unknown = set(spec) - allowed
    if unknown:
        raise ValueError(f"font 包含未知字段：{sorted(unknown)}")
    if "name" in spec:
        font.name = str(spec["name"])
    if "size" in spec:
        size = float(spec["size"])
        if size < 1 or size > 200:
            raise ValueError("字体大小必须在 1 到 200 之间")
        font.sz = size
    for source, target in (
        ("bold", "b"),
        ("italic", "i"),
        ("strike", "strike"),
    ):
        if source in spec:
            setattr(font, target, bool(spec[source]))
    if "underline" in spec:
        underline = spec["underline"]
        if underline is True:
            underline = "single"
        elif underline is False:
            underline = None
        if underline not in {None, "single", "double", "singleAccounting", "doubleAccounting"}:
            raise ValueError("underline 值无效")
        font.u = underline
    if "color" in spec:
        font.color = _color(spec["color"], "font.color")
    if "vert_align" in spec:
        value = spec["vert_align"]
        if value not in {None, "baseline", "superscript", "subscript"}:
            raise ValueError("font.vert_align 值无效")
        font.vertAlign = value
    cell.font = font


def _apply_alignment(cell: Any, spec: dict[str, Any]) -> None:
    alignment = copy(cell.alignment)
    mapping = {
        "horizontal": "horizontal",
        "vertical": "vertical",
        "wrap_text": "wrap_text",
        "shrink_to_fit": "shrink_to_fit",
        "text_rotation": "text_rotation",
        "indent": "indent",
    }
    unknown = set(spec) - set(mapping)
    if unknown:
        raise ValueError(f"alignment 包含未知字段：{sorted(unknown)}")
    for source, target in mapping.items():
        if source in spec:
            setattr(alignment, target, spec[source])
    cell.alignment = alignment


def _apply_border(cell: Any, spec: dict[str, Any]) -> None:
    from openpyxl.styles import Border, Side

    allowed_sides = {"left", "right", "top", "bottom", "diagonal"}
    unknown = set(spec) - allowed_sides
    if unknown:
        raise ValueError(f"border 包含未知字段：{sorted(unknown)}")
    current = cell.border
    sides: dict[str, Any] = {
        name: copy(getattr(current, name)) for name in allowed_sides
    }
    for name, side_spec_raw in spec.items():
        side_spec = _expect_object(side_spec_raw, f"border.{name}")
        unknown_side = set(side_spec) - {"style", "color"}
        if unknown_side:
            raise ValueError(
                f"border.{name} 包含未知字段：{sorted(unknown_side)}"
            )
        sides[name] = Side(
            style=side_spec.get("style"),
            color=(
                _color(side_spec["color"], f"border.{name}.color")
                if side_spec.get("color")
                else None
            ),
        )
    cell.border = Border(
        left=sides["left"],
        right=sides["right"],
        top=sides["top"],
        bottom=sides["bottom"],
        diagonal=sides["diagonal"],
        diagonalUp=current.diagonalUp,
        diagonalDown=current.diagonalDown,
        outline=current.outline,
        vertical=current.vertical,
        horizontal=current.horizontal,
    )


def _apply_style(cell: Any, raw_spec: Any) -> None:
    from openpyxl.styles import PatternFill, Protection

    spec = _expect_object(raw_spec, "style")
    allowed = {
        "named_style",
        "font",
        "fill",
        "alignment",
        "border",
        "number_format",
        "protection",
    }
    unknown = set(spec) - allowed
    if unknown:
        raise ValueError(f"style 包含未知字段：{sorted(unknown)}")
    if "named_style" in spec:
        cell.style = str(spec["named_style"])
    if "font" in spec:
        _apply_font(cell, _expect_object(spec["font"], "font"))
    if "fill" in spec:
        fill_spec = _expect_object(spec["fill"], "fill")
        unknown_fill = set(fill_spec) - {"color", "pattern"}
        if unknown_fill:
            raise ValueError(f"fill 包含未知字段：{sorted(unknown_fill)}")
        pattern = fill_spec.get("pattern", "solid")
        color = _color(fill_spec.get("color"), "fill.color")
        cell.fill = PatternFill(fill_type=pattern, fgColor=color)
    if "alignment" in spec:
        _apply_alignment(
            cell,
            _expect_object(spec["alignment"], "alignment"),
        )
    if "border" in spec:
        _apply_border(cell, _expect_object(spec["border"], "border"))
    if "number_format" in spec:
        cell.number_format = str(spec["number_format"])
    if "protection" in spec:
        protection = _expect_object(spec["protection"], "protection")
        unknown_protection = set(protection) - {"locked", "hidden"}
        if unknown_protection:
            raise ValueError(
                f"protection 包含未知字段：{sorted(unknown_protection)}"
            )
        cell.protection = Protection(
            locked=bool(protection.get("locked", True)),
            hidden=bool(protection.get("hidden", False)),
        )


def _set_cell(cell: Any, item: dict[str, Any]) -> None:
    from openpyxl.comments import Comment

    if "formula" in item and "value" in item:
        raise ValueError(f"{cell.coordinate} 不能同时设置 formula 和 value")
    if "formula" in item:
        formula = item["formula"]
        if not isinstance(formula, str) or not formula.startswith("="):
            raise ValueError(f"{cell.coordinate} 的 formula 必须以 = 开头")
        cell.value = formula
    elif "value" in item:
        cell.value = item["value"]
    if "style" in item:
        _apply_style(cell, item["style"])
    if "comment" in item:
        comment = item["comment"]
        if comment is None:
            cell.comment = None
        elif isinstance(comment, str):
            cell.comment = Comment(comment, "AI")
        else:
            comment_spec = _expect_object(comment, "comment")
            text = str(comment_spec.get("text", ""))
            author = str(comment_spec.get("author", "AI"))
            if not text:
                raise ValueError("comment.text 不能为空")
            cell.comment = Comment(text, author)
    if "hyperlink" in item:
        cell.hyperlink = item["hyperlink"]


def _apply_properties(workbook: Any, raw: Any) -> None:
    spec = _expect_object(raw, "properties")
    allowed = {
        "title",
        "subject",
        "creator",
        "last_modified_by",
        "category",
        "keywords",
        "description",
        "language",
    }
    unknown = set(spec) - allowed
    if unknown:
        raise ValueError(f"properties 包含未知字段：{sorted(unknown)}")
    mapping = {
        "last_modified_by": "lastModifiedBy",
    }
    for key, value in spec.items():
        setattr(workbook.properties, mapping.get(key, key), value)


def _op_add_sheet(workbook: Any, op: dict[str, Any]) -> int:
    name = str(op.get("name", "")).strip()
    if not name:
        raise ValueError("add_sheet.name 不能为空")
    if name in workbook.sheetnames:
        raise ValueError(f"工作表已存在：{name}")
    index = op.get("index")
    if index is not None:
        index = int(index)
        if index < 0 or index > len(workbook.worksheets):
            raise ValueError("add_sheet.index 超出范围")
    workbook.create_sheet(title=name, index=index)
    return 0


def _op_remove_sheet(workbook: Any, op: dict[str, Any]) -> int:
    if len(workbook.worksheets) <= 1:
        raise ValueError("工作簿必须至少保留一个工作表")
    worksheet = _sheet(workbook, op.get("sheet"))
    workbook.remove(worksheet)
    return 0


def _op_rename_sheet(workbook: Any, op: dict[str, Any]) -> int:
    worksheet = _sheet(workbook, op.get("sheet"))
    new_name = str(op.get("name", "")).strip()
    if not new_name:
        raise ValueError("rename_sheet.name 不能为空")
    if new_name in workbook.sheetnames and new_name != worksheet.title:
        raise ValueError(f"工作表已存在：{new_name}")
    worksheet.title = new_name
    return 0


def _op_set_cells(workbook: Any, op: dict[str, Any]) -> int:
    worksheet = _sheet(workbook, op.get("sheet"))
    items = _expect_list(op.get("cells"), "set_cells.cells")
    if len(items) > MAX_WRITTEN_CELLS:
        raise ValueError("单次 set_cells 超过单元格数量限制")
    for raw_item in items:
        item = _expect_object(raw_item, "set_cells.cells[]")
        unknown = set(item) - {
            "cell",
            "value",
            "formula",
            "style",
            "comment",
            "hyperlink",
        }
        if unknown:
            raise ValueError(f"set_cells 单元格包含未知字段：{sorted(unknown)}")
        reference = validate_cell_reference(str(item.get("cell", "")))
        _set_cell(worksheet[reference], item)
    return len(items)


def _op_write_rows(workbook: Any, op: dict[str, Any]) -> int:
    from openpyxl.utils.cell import coordinate_to_tuple

    worksheet = _sheet(workbook, op.get("sheet"))
    start_cell = validate_cell_reference(str(op.get("start_cell", "A1")))
    start_row, start_column = coordinate_to_tuple(start_cell)
    rows = _expect_list(op.get("rows"), "write_rows.rows")
    cell_count = sum(len(row) if isinstance(row, list) else 0 for row in rows)
    if cell_count > MAX_WRITTEN_CELLS:
        raise ValueError("单次 write_rows 超过单元格数量限制")
    for row_offset, raw_row in enumerate(rows):
        row = _expect_list(raw_row, f"write_rows.rows[{row_offset}]")
        for column_offset, raw_value in enumerate(row):
            cell = worksheet.cell(
                row=start_row + row_offset,
                column=start_column + column_offset,
            )
            if isinstance(raw_value, dict) and any(
                key in raw_value
                for key in ("value", "formula", "style", "comment", "hyperlink")
            ):
                _set_cell(cell, raw_value)
            else:
                cell.value = raw_value
    if "style" in op:
        style = op["style"]
        for row_offset, raw_row in enumerate(rows):
            for column_offset in range(len(raw_row)):
                _apply_style(
                    worksheet.cell(
                        row=start_row + row_offset,
                        column=start_column + column_offset,
                    ),
                    style,
                )
    return cell_count


def _op_append_rows(workbook: Any, op: dict[str, Any]) -> int:
    worksheet = _sheet(workbook, op.get("sheet"))
    rows = _expect_list(op.get("rows"), "append_rows.rows")
    cell_count = sum(len(row) if isinstance(row, list) else 0 for row in rows)
    if cell_count > MAX_WRITTEN_CELLS:
        raise ValueError("单次 append_rows 超过单元格数量限制")
    for index, raw_row in enumerate(rows):
        row = _expect_list(raw_row, f"append_rows.rows[{index}]")
        worksheet.append(row)
    return cell_count


def _op_style_range(workbook: Any, op: dict[str, Any]) -> int:
    worksheet = _sheet(workbook, op.get("sheet"))
    reference = validate_cell_range(str(op.get("range", "")))
    cells = list(_iter_range_cells(worksheet, reference))
    if len(cells) > MAX_WRITTEN_CELLS:
        raise ValueError("style_range 超过单元格数量限制")
    for cell in cells:
        _apply_style(cell, op.get("style"))
    return len(cells)


def _op_clear_range(workbook: Any, op: dict[str, Any]) -> int:
    worksheet = _sheet(workbook, op.get("sheet"))
    reference = validate_cell_range(str(op.get("range", "")))
    cells = list(_iter_range_cells(worksheet, reference))
    if len(cells) > MAX_WRITTEN_CELLS:
        raise ValueError("clear_range 超过单元格数量限制")
    clear_values = bool(op.get("values", True))
    clear_styles = bool(op.get("styles", False))
    clear_comments = bool(op.get("comments", False))
    clear_hyperlinks = bool(op.get("hyperlinks", False))
    for cell in cells:
        if clear_values:
            cell.value = None
        if clear_styles:
            cell._style = copy(worksheet.parent._named_styles["Normal"]._style)
        if clear_comments:
            cell.comment = None
        if clear_hyperlinks:
            cell.hyperlink = None
    return len(cells)


def _op_rows_or_columns(workbook: Any, op: dict[str, Any], kind: str) -> int:
    worksheet = _sheet(workbook, op.get("sheet"))
    index = int(op.get("index", 0))
    amount = int(op.get("amount", 1))
    if index < 1:
        raise ValueError(f"{kind}.index 必须大于 0")
    if amount < 1 or amount > 10_000:
        raise ValueError(f"{kind}.amount 必须在 1 到 10000 之间")
    if kind == "insert_rows":
        worksheet.insert_rows(index, amount)
    elif kind == "delete_rows":
        worksheet.delete_rows(index, amount)
    elif kind == "insert_columns":
        worksheet.insert_cols(index, amount)
    elif kind == "delete_columns":
        worksheet.delete_cols(index, amount)
    return 0


def _op_merge(workbook: Any, op: dict[str, Any], merge: bool) -> int:
    worksheet = _sheet(workbook, op.get("sheet"))
    reference = validate_cell_range(str(op.get("range", "")))
    if ":" not in reference:
        raise ValueError("合并区域必须包含起止单元格")
    if merge:
        worksheet.merge_cells(reference)
    else:
        worksheet.unmerge_cells(reference)
    return 0


def _op_set_column_widths(workbook: Any, op: dict[str, Any]) -> int:
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.cell import column_index_from_string

    worksheet = _sheet(workbook, op.get("sheet"))
    widths = _expect_object(op.get("widths"), "set_column_widths.widths")
    changed = 0
    for raw_range, raw_width in widths.items():
        width = float(raw_width)
        if width < 0 or width > 255:
            raise ValueError("列宽必须在 0 到 255 之间")
        pieces = str(raw_range).upper().split(":", 1)
        start = column_index_from_string(pieces[0])
        end = column_index_from_string(pieces[-1])
        if start > end:
            raise ValueError(f"列范围无效：{raw_range}")
        for column in range(start, end + 1):
            worksheet.column_dimensions[get_column_letter(column)].width = width
            changed += 1
    return changed


def _op_set_row_heights(workbook: Any, op: dict[str, Any]) -> int:
    worksheet = _sheet(workbook, op.get("sheet"))
    heights = _expect_object(op.get("heights"), "set_row_heights.heights")
    changed = 0
    for raw_range, raw_height in heights.items():
        height = float(raw_height)
        if height < 0 or height > 409:
            raise ValueError("行高必须在 0 到 409 之间")
        pieces = str(raw_range).split(":", 1)
        start = int(pieces[0])
        end = int(pieces[-1])
        if start < 1 or start > end or end > 1_048_576:
            raise ValueError(f"行范围无效：{raw_range}")
        for row in range(start, end + 1):
            worksheet.row_dimensions[row].height = height
            changed += 1
    return changed


def _op_add_table(workbook: Any, op: dict[str, Any]) -> int:
    from openpyxl.worksheet.table import Table, TableStyleInfo

    worksheet = _sheet(workbook, op.get("sheet"))
    reference = validate_cell_range(str(op.get("range", "")))
    if ":" not in reference:
        raise ValueError("表格区域必须包含多于一个单元格")
    name = str(op.get("name", "")).strip()
    if not name or not TABLE_NAME_RE.fullmatch(name):
        raise ValueError("add_table.name 必须是合法的 Excel 表格名称")
    if any(name in sheet.tables for sheet in workbook.worksheets):
        raise ValueError(f"表格名称已存在：{name}")
    table = Table(displayName=name, ref=reference)
    table.tableStyleInfo = TableStyleInfo(
        name=str(op.get("style", "TableStyleMedium2")),
        showFirstColumn=bool(op.get("show_first_column", False)),
        showLastColumn=bool(op.get("show_last_column", False)),
        showRowStripes=bool(op.get("show_row_stripes", True)),
        showColumnStripes=bool(op.get("show_column_stripes", False)),
    )
    worksheet.add_table(table)
    return 0


def _op_add_chart(workbook: Any, op: dict[str, Any]) -> int:
    from openpyxl.chart import AreaChart, BarChart, LineChart, PieChart, Reference
    from openpyxl.utils.cell import range_boundaries

    worksheet = _sheet(workbook, op.get("sheet"))
    chart_type = str(op.get("chart_type", "bar")).lower()
    chart_classes = {
        "area": AreaChart,
        "bar": BarChart,
        "column": BarChart,
        "line": LineChart,
        "pie": PieChart,
    }
    if chart_type not in chart_classes:
        raise ValueError("chart_type 仅支持 area、bar、column、line、pie")
    data_range = validate_cell_range(str(op.get("data_range", "")))
    min_col, min_row, max_col, max_row = range_boundaries(data_range)
    chart = chart_classes[chart_type]()
    if isinstance(chart, BarChart):
        chart.type = "bar" if chart_type == "bar" else "col"
    data = Reference(
        worksheet,
        min_col=min_col,
        min_row=min_row,
        max_col=max_col,
        max_row=max_row,
    )
    chart.add_data(
        data,
        titles_from_data=bool(op.get("titles_from_data", True)),
        from_rows=bool(op.get("from_rows", False)),
    )
    if op.get("categories_range"):
        category_range = validate_cell_range(str(op["categories_range"]))
        c_min_col, c_min_row, c_max_col, c_max_row = range_boundaries(
            category_range
        )
        categories = Reference(
            worksheet,
            min_col=c_min_col,
            min_row=c_min_row,
            max_col=c_max_col,
            max_row=c_max_row,
        )
        chart.set_categories(categories)
    if "title" in op:
        chart.title = str(op["title"])
    if "x_axis_title" in op and hasattr(chart, "x_axis"):
        chart.x_axis.title = str(op["x_axis_title"])
    if "y_axis_title" in op and hasattr(chart, "y_axis"):
        chart.y_axis.title = str(op["y_axis_title"])
    if "style" in op:
        chart.style = int(op["style"])
    if "height" in op:
        chart.height = float(op["height"])
    if "width" in op:
        chart.width = float(op["width"])
    if "legend_position" in op and chart.legend:
        chart.legend.position = str(op["legend_position"])
    anchor = validate_cell_reference(str(op.get("anchor", "E2")))
    worksheet.add_chart(chart, anchor)
    return 0


def _op_add_image(workbook: Any, op: dict[str, Any]) -> int:
    from openpyxl.drawing.image import Image

    worksheet = _sheet(workbook, op.get("sheet"))
    image_path = input_file(
        str(op.get("path", "")),
        {".png", ".jpg", ".jpeg", ".gif", ".bmp"},
    )
    image = Image(str(image_path))
    if "width" in op:
        image.width = float(op["width"])
    if "height" in op:
        image.height = float(op["height"])
    anchor = validate_cell_reference(str(op.get("anchor", "A1")))
    worksheet.add_image(image, anchor)
    return 0


def _op_add_data_validation(workbook: Any, op: dict[str, Any]) -> int:
    from openpyxl.worksheet.datavalidation import DataValidation

    worksheet = _sheet(workbook, op.get("sheet"))
    reference = validate_cell_range(str(op.get("range", "")))
    validation_type = str(op.get("validation_type", "list"))
    allowed = {
        "list",
        "whole",
        "decimal",
        "date",
        "time",
        "textLength",
        "custom",
    }
    if validation_type not in allowed:
        raise ValueError(f"validation_type 不支持：{validation_type}")
    validation = DataValidation(
        type=validation_type,
        formula1=op.get("formula1"),
        formula2=op.get("formula2"),
        operator=op.get("operator"),
        allow_blank=bool(op.get("allow_blank", True)),
        showErrorMessage=bool(op.get("show_error_message", True)),
        errorTitle=op.get("error_title"),
        error=op.get("error_message"),
        promptTitle=op.get("prompt_title"),
        prompt=op.get("prompt_message"),
        showInputMessage=bool(op.get("show_input_message", True)),
    )
    worksheet.add_data_validation(validation)
    validation.add(reference)
    return 0


def _op_add_conditional_format(workbook: Any, op: dict[str, Any]) -> int:
    from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
    from openpyxl.styles import Font, PatternFill

    worksheet = _sheet(workbook, op.get("sheet"))
    reference = validate_cell_range(str(op.get("range", "")))
    rule_type = str(op.get("rule_type", "cell"))
    if rule_type == "color_scale":
        colors = op.get("colors", ["F8696B", "FFEB84", "63BE7B"])
        if not isinstance(colors, list) or len(colors) not in {2, 3}:
            raise ValueError("color_scale.colors 必须包含 2 或 3 个颜色")
        colors = [_color(item, "colors[]") for item in colors]
        if len(colors) == 2:
            rule = ColorScaleRule(
                start_type="min",
                start_color=colors[0],
                end_type="max",
                end_color=colors[1],
            )
        else:
            rule = ColorScaleRule(
                start_type="min",
                start_color=colors[0],
                mid_type="percentile",
                mid_value=50,
                mid_color=colors[1],
                end_type="max",
                end_color=colors[2],
            )
    else:
        formulas = op.get("formula")
        if isinstance(formulas, str):
            formulas = [formulas]
        formulas = _expect_list(formulas, "conditional_format.formula")
        fill = (
            PatternFill(
                fill_type="solid",
                fgColor=_color(op["fill_color"], "fill_color"),
            )
            if op.get("fill_color")
            else None
        )
        font = (
            Font(color=_color(op["font_color"], "font_color"))
            if op.get("font_color")
            else None
        )
        if rule_type == "cell":
            rule = CellIsRule(
                operator=str(op.get("operator", "equal")),
                formula=formulas,
                fill=fill,
                font=font,
                stopIfTrue=bool(op.get("stop_if_true", False)),
            )
        elif rule_type == "formula":
            rule = FormulaRule(
                formula=formulas,
                fill=fill,
                font=font,
                stopIfTrue=bool(op.get("stop_if_true", False)),
            )
        else:
            raise ValueError("rule_type 仅支持 cell、formula、color_scale")
    worksheet.conditional_formatting.add(reference, rule)
    return 0


def _op_set_print(workbook: Any, op: dict[str, Any]) -> int:
    worksheet = _sheet(workbook, op.get("sheet"))
    if "print_area" in op:
        worksheet.print_area = validate_cell_range(str(op["print_area"]))
    if "orientation" in op:
        orientation = str(op["orientation"]).lower()
        if orientation not in {"portrait", "landscape"}:
            raise ValueError("orientation 只能是 portrait 或 landscape")
        worksheet.page_setup.orientation = orientation
    if "paper_size" in op:
        paper = str(op["paper_size"]).upper()
        paper_map = {
            "A4": worksheet.PAPERSIZE_A4,
            "LETTER": worksheet.PAPERSIZE_LETTER,
            "LEGAL": worksheet.PAPERSIZE_LEGAL,
        }
        if paper not in paper_map:
            raise ValueError("paper_size 仅支持 A4、LETTER、LEGAL")
        worksheet.page_setup.paperSize = paper_map[paper]
    if "fit_to_width" in op:
        worksheet.page_setup.fitToWidth = int(op["fit_to_width"])
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    if "fit_to_height" in op:
        worksheet.page_setup.fitToHeight = int(op["fit_to_height"])
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    if "repeat_rows" in op:
        worksheet.print_title_rows = str(op["repeat_rows"])
    if "repeat_columns" in op:
        worksheet.print_title_cols = str(op["repeat_columns"])
    margins = op.get("margins")
    if margins is not None:
        margins_spec = _expect_object(margins, "set_print.margins")
        for key in ("left", "right", "top", "bottom", "header", "footer"):
            if key in margins_spec:
                setattr(worksheet.page_margins, key, float(margins_spec[key]))
    if "show_gridlines" in op:
        worksheet.sheet_view.showGridLines = bool(op["show_gridlines"])
    if "center_horizontally" in op:
        worksheet.print_options.horizontalCentered = bool(
            op["center_horizontally"]
        )
    return 0


def _op_set_named_range(workbook: Any, op: dict[str, Any]) -> int:
    from openpyxl.workbook.defined_name import DefinedName

    worksheet = _sheet(workbook, op.get("sheet"))
    name = str(op.get("name", "")).strip()
    if not name or not TABLE_NAME_RE.fullmatch(name):
        raise ValueError("set_named_range.name 无效")
    reference = validate_cell_range(str(op.get("range", "")))
    quoted_sheet = worksheet.title.replace("'", "''")
    attr_text = f"'{quoted_sheet}'!{reference}"
    defined = DefinedName(name, attr_text=attr_text)
    try:
        workbook.defined_names.add(defined)
    except AttributeError:
        workbook.defined_names.append(defined)
    return 0


def _op_freeze_panes(workbook: Any, op: dict[str, Any]) -> int:
    worksheet = _sheet(workbook, op.get("sheet"))
    value = op.get("cell")
    worksheet.freeze_panes = (
        validate_cell_reference(str(value)) if value not in {None, ""} else None
    )
    return 0


def _op_auto_filter(workbook: Any, op: dict[str, Any]) -> int:
    worksheet = _sheet(workbook, op.get("sheet"))
    value = op.get("range")
    worksheet.auto_filter.ref = (
        validate_cell_range(str(value)) if value not in {None, ""} else None
    )
    return 0


def _apply_operation(workbook: Any, raw_op: Any) -> int:
    op = _expect_object(raw_op, "operations[]")
    kind = str(op.get("type", "")).strip()
    handlers = {
        "add_sheet": _op_add_sheet,
        "remove_sheet": _op_remove_sheet,
        "rename_sheet": _op_rename_sheet,
        "set_cells": _op_set_cells,
        "write_rows": _op_write_rows,
        "append_rows": _op_append_rows,
        "style_range": _op_style_range,
        "clear_range": _op_clear_range,
        "set_column_widths": _op_set_column_widths,
        "set_row_heights": _op_set_row_heights,
        "add_table": _op_add_table,
        "add_chart": _op_add_chart,
        "add_image": _op_add_image,
        "add_data_validation": _op_add_data_validation,
        "add_conditional_format": _op_add_conditional_format,
        "set_print": _op_set_print,
        "set_named_range": _op_set_named_range,
        "freeze_panes": _op_freeze_panes,
        "set_auto_filter": _op_auto_filter,
    }
    if kind in {"insert_rows", "delete_rows", "insert_columns", "delete_columns"}:
        return _op_rows_or_columns(workbook, op, kind)
    if kind == "merge_cells":
        return _op_merge(workbook, op, True)
    if kind == "unmerge_cells":
        return _op_merge(workbook, op, False)
    handler = handlers.get(kind)
    if handler is None:
        supported = sorted(
            set(handlers)
            | {
                "insert_rows",
                "delete_rows",
                "insert_columns",
                "delete_columns",
                "merge_cells",
                "unmerge_cells",
            }
        )
        raise ValueError(f"不支持的操作 {kind!r}；可选：{'、'.join(supported)}")
    return handler(workbook, op)


def _scan_workbook(workbook: Any) -> dict[str, Any]:
    formula_count = 0
    literal_errors: list[dict[str, str]] = []
    risky_formulas: list[dict[str, str]] = []
    for worksheet in workbook.worksheets:
        for cell in worksheet._cells.values():
            value = cell.value
            if cell.data_type == "f" or (
                isinstance(value, str) and value.startswith("=")
            ):
                formula_count += 1
                if (
                    isinstance(value, str)
                    and RISKY_FORMULA_RE.search(value)
                    and len(risky_formulas) < 100
                ):
                    risky_formulas.append(
                        {
                            "sheet": worksheet.title,
                            "cell": cell.coordinate,
                            "formula": value,
                        }
                    )
            error = normalize_formula_error(value)
            if error and len(literal_errors) < 100:
                literal_errors.append(
                    {
                        "sheet": worksheet.title,
                        "cell": cell.coordinate,
                        "error": error,
                    }
                )
    return {
        "formula_count": formula_count,
        "literal_errors": literal_errors,
        "risky_formulas": risky_formulas,
    }


def build_parser() -> argparse.ArgumentParser:
    parser = SkillArgumentParser(
        description="依据受控 JSON 操作创建或编辑 Excel 工作簿。"
    )
    parser.add_argument("--input", help="可选的现有 .xlsx/.xlsm/.xltx 文件")
    parser.add_argument("--output", required=True, help="输出 .xlsx 或 .xlsm")
    parser.add_argument("--spec", help="内联 JSON 操作说明")
    parser.add_argument("--spec-file", help="JSON 操作说明文件")
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument(
        "--allow-external-links",
        action="store_true",
        help="明确接受外部链接缓存值可能丢失的风险",
    )
    parser.add_argument(
        "--drop-macros",
        action="store_true",
        help="明确允许把 .xlsm 输出为 .xlsx 并丢弃宏",
    )
    return parser


def main() -> dict[str, Any]:
    from openpyxl import Workbook, load_workbook

    args = build_parser().parse_args()
    destination = output_file(
        args.output,
        EXCEL_OUTPUT_SUFFIXES,
        overwrite=args.overwrite,
    )
    spec = load_json_argument(args.spec, args.spec_file, label="操作说明")
    unknown_top = set(spec) - {
        "properties",
        "operations",
        "active_sheet",
        "calculation_mode",
    }
    if unknown_top:
        raise ValueError(f"操作说明包含未知顶层字段：{sorted(unknown_top)}")

    source: Optional[Path] = None
    warnings: list[str] = []
    if args.input:
        source = input_file(args.input, EXCEL_INPUT_SUFFIXES)
        if source == destination:
            raise ValueError("输出路径不能与输入文件相同；请保留原始文件")
        has_external_links = workbook_has_external_links(source)
        if has_external_links and not args.allow_external_links:
            raise ValueError(
                "输入工作簿含外部链接。openpyxl 保存后可能丢失缓存值；"
                "请先固化外部数据，或明确传 --allow-external-links 接受风险"
            )
        if has_external_links:
            warnings.append("输入工作簿含外部链接，保存后外部缓存值可能变化")
        source_macro = source.suffix.lower() in {".xlsm", ".xltm"}
        destination_macro = destination.suffix.lower() == ".xlsm"
        if source_macro and not destination_macro and not args.drop_macros:
            raise ValueError(
                "宏工作簿输出为 .xlsx 会丢失宏；如确认接受请传 --drop-macros"
            )
        if not source_macro and destination_macro:
            raise ValueError("不能从无宏工作簿新建真正的 .xlsm 文件")
        if source_macro and not destination_macro:
            warnings.append("已按要求把宏工作簿输出为 .xlsx，宏将被丢弃")
        workbook = load_workbook(source, **openpyxl_load_options(source))
    else:
        if destination.suffix.lower() == ".xlsm":
            raise ValueError("新建工作簿只支持 .xlsx；无 VBA 工程时不能伪造 .xlsm")
        workbook = Workbook()

    if "properties" in spec:
        _apply_properties(workbook, spec["properties"])
    operations = _expect_list(spec.get("operations", []), "operations")
    if len(operations) > MAX_OPERATIONS:
        raise ValueError(f"operations 不能超过 {MAX_OPERATIONS} 项")

    written_cells = 0
    for index, operation in enumerate(operations):
        try:
            written_cells += _apply_operation(workbook, operation)
        except Exception as exc:
            raise ValueError(f"第 {index + 1} 个操作失败：{exc}") from exc
        if written_cells > MAX_WRITTEN_CELLS:
            raise ValueError(
                f"本次任务写入或处理的单元格不能超过 {MAX_WRITTEN_CELLS}"
            )

    if "active_sheet" in spec:
        active_name = str(spec["active_sheet"])
        if active_name not in workbook.sheetnames:
            raise ValueError(f"active_sheet 不存在：{active_name}")
        workbook.active = workbook.sheetnames.index(active_name)

    calculation_mode = str(spec.get("calculation_mode", "auto"))
    if calculation_mode not in {"auto", "manual", "autoNoTable"}:
        raise ValueError("calculation_mode 只能是 auto、manual、autoNoTable")
    if getattr(workbook, "calculation", None):
        workbook.calculation.calcMode = calculation_mode
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True

    scan = _scan_workbook(workbook)
    if scan["literal_errors"]:
        warnings.append("工作簿中存在字面错误值，请在交付前检查")
    if scan["risky_formulas"]:
        warnings.append(
            "检测到 LibreOffice 可能无法可靠重算的动态数组/新式查找公式"
        )

    descriptor, temp_name = tempfile.mkstemp(
        prefix=f".{destination.stem}.",
        suffix=destination.suffix,
        dir=str(destination.parent),
    )
    os.close(descriptor)
    temp_path = Path(temp_name)
    try:
        workbook.save(temp_path)
        load_workbook(
            temp_path,
            read_only=True,
            data_only=False,
            keep_vba=destination.suffix.lower() == ".xlsm",
            keep_links=True,
        ).close()
        publish_file(temp_path, destination, overwrite=args.overwrite)
    finally:
        if temp_path.exists():
            temp_path.unlink()

    return {
        "path": str(destination),
        "source": str(source) if source else None,
        "sheet_names": workbook.sheetnames,
        "active_sheet": workbook.active.title,
        "operation_count": len(operations),
        "processed_cell_count": written_cells,
        "formula_count": scan["formula_count"],
        "requires_recalculation": scan["formula_count"] > 0,
        "literal_errors": scan["literal_errors"],
        "risky_formulas": scan["risky_formulas"],
        "warnings": warnings,
    }


if __name__ == "__main__":
    raise SystemExit(run_cli(main))
