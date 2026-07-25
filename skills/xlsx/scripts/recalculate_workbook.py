#!/usr/bin/env python3

from __future__ import annotations

import argparse
import shutil
import tempfile
from collections import defaultdict
from pathlib import Path
from typing import Any

from _xlsx_common import (
    EXCEL_OUTPUT_SUFFIXES,
    SkillArgumentParser,
    input_file,
    normalize_formula_error,
    openpyxl_load_options,
    output_file,
    publish_file,
    run_cli,
    run_soffice_convert,
    workbook_has_external_links,
)


def _scan_formula_results(path: Path) -> dict[str, Any]:
    from openpyxl import load_workbook

    options = openpyxl_load_options(path)
    formulas = load_workbook(path, data_only=False, **options)
    cached = load_workbook(path, data_only=True, **options)
    error_locations: dict[str, list[dict[str, str]]] = defaultdict(list)
    total_errors_by_type: dict[str, int] = defaultdict(int)
    total_formulas = 0
    missing_cached = 0

    for formula_sheet in formulas.worksheets:
        cached_sheet = cached[formula_sheet.title]
        for cell in formula_sheet._cells.values():
            is_formula = cell.data_type == "f" or (
                isinstance(cell.value, str) and cell.value.startswith("=")
            )
            if not is_formula:
                continue
            total_formulas += 1
            cached_value = cached_sheet[cell.coordinate].value
            error = normalize_formula_error(cached_value)
            if error:
                total_errors_by_type[error] += 1
                if len(error_locations[error]) < 100:
                    error_locations[error].append(
                        {
                            "sheet": formula_sheet.title,
                            "cell": cell.coordinate,
                            "formula": str(cell.value),
                        }
                    )
            elif cached_value is None:
                missing_cached += 1

    summaries = []
    for error_type in sorted(total_errors_by_type):
        total = total_errors_by_type[error_type]
        locations = error_locations[error_type]
        summaries.append(
            {
                "error": error_type,
                "count": total,
                "locations": locations,
                "locations_truncated": max(0, total - len(locations)),
            }
        )
    formulas.close()
    cached.close()
    return {
        "total_formulas": total_formulas,
        "total_errors": sum(total_errors_by_type.values()),
        "error_summary": summaries,
        "missing_cached_value_count": missing_cached,
    }


def build_parser() -> argparse.ArgumentParser:
    parser = SkillArgumentParser(
        description="用 LibreOffice 重算工作簿，并检查公式缓存错误。"
    )
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--timeout", type=int, default=120)
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument(
        "--allow-external-links",
        action="store_true",
        help="明确接受外部链接可能失效或缓存变化的风险",
    )
    return parser


def main() -> dict[str, Any]:
    args = build_parser().parse_args()
    source = input_file(args.input, EXCEL_OUTPUT_SUFFIXES)
    destination = output_file(
        args.output,
        EXCEL_OUTPUT_SUFFIXES,
        overwrite=args.overwrite,
    )
    if source == destination:
        raise ValueError("重算结果必须写到新文件，不能覆盖输入源文件")
    if source.suffix.lower() != destination.suffix.lower():
        raise ValueError("重算必须保持原格式（.xlsx 或 .xlsm），不能同时转换格式")
    has_external_links = workbook_has_external_links(source)
    if has_external_links and not args.allow_external_links:
        raise ValueError(
            "工作簿包含外部链接，重算可能把缓存值改为 #NAME? 并删除链接；"
            "请先固化外部数据，或明确传 --allow-external-links"
        )

    with tempfile.TemporaryDirectory(prefix="xlsx-recalculate-") as temp_name:
        temp_dir = Path(temp_name)
        input_dir = temp_dir / "input"
        output_dir = temp_dir / "output"
        input_dir.mkdir()
        output_dir.mkdir()
        staged_source = input_dir / f"source{source.suffix.lower()}"
        shutil.copy2(source, staged_source)
        target_format = source.suffix.lower().lstrip(".")
        filter_name = (
            "Calc MS Excel 2007 VBA XML"
            if target_format == "xlsm"
            else "Calc MS Excel 2007 XML"
        )
        converted, office_output = run_soffice_convert(
            staged_source,
            target_format=target_format,
            output_dir=output_dir,
            timeout=args.timeout,
            filter_name=filter_name,
        )
        scan = _scan_formula_results(converted)
        staged_result = temp_dir / f"result{destination.suffix.lower()}"
        shutil.copy2(converted, staged_result)
        publish_file(staged_result, destination, overwrite=args.overwrite)

    warnings: list[str] = []
    if has_external_links:
        warnings.append("工作簿含外部链接，已按明确授权执行重算")
    if scan["missing_cached_value_count"]:
        warnings.append(
            "部分公式缓存值为空；公式结果可能确实为空字符串，也可能需要人工核验"
        )
    return {
        "path": str(destination),
        "source": str(source),
        "status": "success" if scan["total_errors"] == 0 else "errors_found",
        **scan,
        "has_external_links": has_external_links,
        "warnings": warnings,
        "office_stdout": office_output["stdout"],
        "office_stderr": office_output["stderr"],
    }


if __name__ == "__main__":
    raise SystemExit(run_cli(main))
