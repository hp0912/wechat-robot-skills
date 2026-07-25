#!/usr/bin/env python3

from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Any, Optional

from _xlsx_common import (
    EXCEL_INPUT_SUFFIXES,
    TABULAR_INPUT_SUFFIXES,
    SkillArgumentParser,
    input_file,
    normalize_formula_error,
    openpyxl_load_options,
    run_cli,
    workbook_has_external_links,
)


def _color_value(color: Any) -> Optional[str]:
    if color is None:
        return None
    color_type = getattr(color, "type", None)
    if color_type == "rgb":
        return getattr(color, "rgb", None)
    if color_type == "theme":
        theme = getattr(color, "theme", None)
        tint = getattr(color, "tint", 0)
        return f"theme:{theme}:tint:{tint}"
    if color_type == "indexed":
        return f"indexed:{getattr(color, 'indexed', None)}"
    return None


def _cell_payload(formula_cell: Any, cached_cell: Any) -> Optional[dict[str, Any]]:
    formula_value = formula_cell.value
    cached_value = cached_cell.value
    if formula_value is None and cached_value is None and not formula_cell.has_style:
        return None

    is_formula = formula_cell.data_type == "f" or (
        isinstance(formula_value, str) and formula_value.startswith("=")
    )
    value = cached_value if is_formula else formula_value
    error = normalize_formula_error(value)
    payload: dict[str, Any] = {
        "cell": formula_cell.coordinate,
        "value": value,
        "type": formula_cell.data_type,
    }
    if is_formula:
        payload["formula"] = formula_value
        payload["cached_value"] = cached_value
    if error:
        payload["error"] = error
    if formula_cell.has_style:
        payload["style"] = {
            "style_id": formula_cell.style_id,
            "number_format": formula_cell.number_format,
            "font": formula_cell.font.name,
            "font_size": formula_cell.font.sz,
            "bold": bool(formula_cell.font.bold),
            "italic": bool(formula_cell.font.italic),
            "font_color": _color_value(formula_cell.font.color),
            "fill_color": _color_value(formula_cell.fill.fgColor),
            "horizontal": formula_cell.alignment.horizontal,
            "vertical": formula_cell.alignment.vertical,
            "wrap_text": bool(formula_cell.alignment.wrap_text),
            "locked": bool(formula_cell.protection.locked),
        }
    if formula_cell.comment:
        payload["comment"] = {
            "author": formula_cell.comment.author,
            "text": formula_cell.comment.text,
        }
    if formula_cell.hyperlink:
        payload["hyperlink"] = formula_cell.hyperlink.target
    return payload


def _defined_names(workbook: Any) -> list[dict[str, Any]]:
    names: list[dict[str, Any]] = []
    try:
        values = workbook.defined_names.values()
    except AttributeError:
        values = workbook.defined_names.definedName
    for item in values:
        names.append(
            {
                "name": getattr(item, "name", None),
                "value": getattr(item, "attr_text", None),
                "local_sheet_id": getattr(item, "localSheetId", None),
                "hidden": bool(getattr(item, "hidden", False)),
            }
        )
    return names


def inspect_excel(
    source: Path,
    *,
    sheet_name: Optional[str],
    start_row: int,
    start_column: int,
    max_rows: int,
    max_columns: int,
) -> dict[str, Any]:
    from openpyxl import load_workbook

    options = openpyxl_load_options(source)
    formulas = load_workbook(source, data_only=False, **options)
    cached = load_workbook(source, data_only=True, **options)

    summaries: list[dict[str, Any]] = []
    total_formulas = 0
    total_errors = 0
    for worksheet in formulas.worksheets:
        formula_count = 0
        error_count = 0
        for cell in worksheet._cells.values():
            if cell.data_type == "f" or (
                isinstance(cell.value, str) and cell.value.startswith("=")
            ):
                formula_count += 1
            if normalize_formula_error(cell.value):
                error_count += 1
        total_formulas += formula_count
        total_errors += error_count
        summaries.append(
            {
                "name": worksheet.title,
                "state": worksheet.sheet_state,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "freeze_panes": (
                    str(worksheet.freeze_panes) if worksheet.freeze_panes else None
                ),
                "auto_filter": worksheet.auto_filter.ref,
                "merged_ranges": [str(item) for item in worksheet.merged_cells.ranges],
                "tables": list(worksheet.tables.keys()),
                "chart_count": len(worksheet._charts),
                "image_count": len(worksheet._images),
                "formula_count": formula_count,
                "literal_error_count": error_count,
                "print_area": str(worksheet.print_area) if worksheet.print_area else None,
            }
        )

    if sheet_name:
        if sheet_name not in formulas.sheetnames:
            raise ValueError(
                f"工作表不存在：{sheet_name}；可选：{'、'.join(formulas.sheetnames)}"
            )
        selected_name = sheet_name
    else:
        selected_name = formulas.active.title

    formula_sheet = formulas[selected_name]
    cached_sheet = cached[selected_name]
    end_row = min(formula_sheet.max_row, start_row + max_rows - 1)
    end_column = min(
        formula_sheet.max_column,
        start_column + max_columns - 1,
    )
    cells: list[dict[str, Any]] = []
    for row in formula_sheet.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=start_column,
        max_col=end_column,
    ):
        for formula_cell in row:
            cached_cell = cached_sheet[formula_cell.coordinate]
            payload = _cell_payload(formula_cell, cached_cell)
            if payload:
                cells.append(payload)

    next_row = end_row + 1 if end_row < formula_sheet.max_row else None
    next_column = (
        end_column + 1 if end_column < formula_sheet.max_column else None
    )
    properties = formulas.properties
    calculation = getattr(formulas, "calculation", None)
    return {
        "path": str(source),
        "format": source.suffix.lower(),
        "macro_enabled": source.suffix.lower() in {".xlsm", ".xltm"},
        "has_external_links": workbook_has_external_links(source),
        "active_sheet": formulas.active.title,
        "sheet_names": formulas.sheetnames,
        "sheets": summaries,
        "defined_names": _defined_names(formulas),
        "properties": {
            "title": properties.title,
            "subject": properties.subject,
            "creator": properties.creator,
            "last_modified_by": properties.lastModifiedBy,
            "created": properties.created,
            "modified": properties.modified,
            "category": properties.category,
            "keywords": properties.keywords,
            "description": properties.description,
        },
        "calculation": {
            "mode": getattr(calculation, "calcMode", None),
            "full_calc_on_load": getattr(calculation, "fullCalcOnLoad", None),
            "force_full_calc": getattr(calculation, "forceFullCalc", None),
        },
        "formula_count": total_formulas,
        "literal_error_count": total_errors,
        "selection": {
            "sheet": selected_name,
            "start_row": start_row,
            "end_row": end_row,
            "start_column": start_column,
            "end_column": end_column,
            "cells": cells,
            "has_more": next_row is not None or next_column is not None,
            "next_row": next_row,
            "next_column": next_column,
        },
    }


def inspect_delimited(
    source: Path,
    *,
    start_row: int,
    start_column: int,
    max_rows: int,
    max_columns: int,
) -> dict[str, Any]:
    delimiter = "\t" if source.suffix.lower() == ".tsv" else ","
    rows: list[list[str]] = []
    total_rows = 0
    max_seen_columns = 0
    with source.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        for index, row in enumerate(reader, start=1):
            total_rows = index
            max_seen_columns = max(max_seen_columns, len(row))
            if index < start_row or len(rows) >= max_rows:
                continue
            rows.append(row[start_column - 1 : start_column - 1 + max_columns])

    return {
        "path": str(source),
        "format": source.suffix.lower(),
        "delimiter": delimiter,
        "row_count": total_rows,
        "max_column_count": max_seen_columns,
        "selection": {
            "start_row": start_row,
            "end_row": min(total_rows, start_row + len(rows) - 1),
            "start_column": start_column,
            "end_column": min(
                max_seen_columns, start_column + max_columns - 1
            ),
            "rows": rows,
            "has_more": (
                start_row + len(rows) - 1 < total_rows
                or start_column + max_columns - 1 < max_seen_columns
            ),
            "next_row": (
                start_row + len(rows)
                if start_row + len(rows) - 1 < total_rows
                else None
            ),
        },
    }


def build_parser() -> argparse.ArgumentParser:
    parser = SkillArgumentParser(
        description="读取 Excel、CSV 或 TSV 的结构、公式、缓存值和局部单元格。"
    )
    parser.add_argument("--input", required=True, help="输入文件路径")
    parser.add_argument("--sheet", help="要读取的工作表；默认活动工作表")
    parser.add_argument("--start-row", type=int, default=1)
    parser.add_argument("--start-column", type=int, default=1)
    parser.add_argument("--max-rows", type=int, default=40)
    parser.add_argument("--max-columns", type=int, default=20)
    return parser


def main() -> dict[str, Any]:
    args = build_parser().parse_args()
    if args.start_row < 1 or args.start_row > 1_048_576:
        raise ValueError("start-row 必须在 1 到 1048576 之间")
    if args.start_column < 1 or args.start_column > 16_384:
        raise ValueError("start-column 必须在 1 到 16384 之间")
    if args.max_rows < 1 or args.max_rows > 200:
        raise ValueError("max-rows 必须在 1 到 200 之间")
    if args.max_columns < 1 or args.max_columns > 100:
        raise ValueError("max-columns 必须在 1 到 100 之间")

    source = input_file(args.input, TABULAR_INPUT_SUFFIXES)
    if source.suffix.lower() == ".xls":
        raise ValueError(
            "旧版 .xls 请先使用 convert_workbook.py 转为 .xlsx 后再检查"
        )
    if source.suffix.lower() in EXCEL_INPUT_SUFFIXES:
        return inspect_excel(
            source,
            sheet_name=args.sheet,
            start_row=args.start_row,
            start_column=args.start_column,
            max_rows=args.max_rows,
            max_columns=args.max_columns,
        )
    if args.sheet:
        raise ValueError("CSV/TSV 没有工作表，不能传 --sheet")
    return inspect_delimited(
        source,
        start_row=args.start_row,
        start_column=args.start_column,
        max_rows=args.max_rows,
        max_columns=args.max_columns,
    )


if __name__ == "__main__":
    raise SystemExit(run_cli(main))
