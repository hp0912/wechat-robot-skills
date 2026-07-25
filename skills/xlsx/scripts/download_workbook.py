#!/usr/bin/env python3

from __future__ import annotations

import argparse
import os
import socket
import stat
import sys
import tempfile
import urllib.error
import urllib.parse
import urllib.request
import zipfile
from pathlib import Path, PurePosixPath
from typing import Any, NoReturn, Optional

from _xlsx_common import (
    TABULAR_INPUT_SUFFIXES,
    emit,
    failure_message,
    output_file,
    publish_file,
)


DEFAULT_TIMEOUT_SECONDS = 60
DEFAULT_MAX_BYTES = 100 * 1024 * 1024
MAX_ALLOWED_BYTES = 512 * 1024 * 1024
CHUNK_SIZE = 1024 * 1024
MAX_ARCHIVE_MEMBERS = 20_000
MAX_ARCHIVE_UNCOMPRESSED_BYTES = 512 * 1024 * 1024
MAX_ARCHIVE_MEMBER_BYTES = 128 * 1024 * 1024
USER_AGENT = "wechat-robot-xlsx-skill/1.0"
OLE_COMPOUND_MAGIC = bytes.fromhex("D0CF11E0A1B11AE1")
CONTENT_TYPES_NS = (
    "http://schemas.openxmlformats.org/package/2006/content-types"
)
WORKBOOK_CONTENT_TYPES = {
    "application/vnd.openxmlformats-officedocument."
    "spreadsheetml.sheet.main+xml": ".xlsx",
    "application/vnd.ms-excel.sheet.macroEnabled.main+xml": ".xlsm",
    "application/vnd.openxmlformats-officedocument."
    "spreadsheetml.template.main+xml": ".xltx",
    "application/vnd.ms-excel.template.macroEnabled.main+xml": ".xltm",
}


class SkillArgumentParser(argparse.ArgumentParser):
    def error(self, message: str) -> NoReturn:
        raise ValueError(f"参数错误：{message}")


def _validate_https_url(value: str) -> str:
    url = value.strip()
    if not url:
        raise ValueError("Excel URL 不能为空")
    if any(character.isspace() or ord(character) < 32 for character in url):
        raise ValueError("Excel URL 不能包含空白字符或控制字符")

    parsed = urllib.parse.urlsplit(url)
    if parsed.scheme.lower() != "https" or not parsed.hostname:
        raise ValueError("Excel URL 必须是有效的 HTTPS 地址")
    if parsed.username is not None or parsed.password is not None:
        raise ValueError("Excel URL 不允许包含用户名或密码")
    try:
        parsed.port
    except ValueError as exc:
        raise ValueError("Excel URL 端口格式不正确") from exc
    return url


class HTTPSOnlyRedirectHandler(urllib.request.HTTPRedirectHandler):
    def redirect_request(self, req, fp, code, msg, headers, newurl):
        safe_url = _validate_https_url(newurl)
        return super().redirect_request(req, fp, code, msg, headers, safe_url)


def _parse_args(argv: list[str]) -> argparse.Namespace:
    parser = SkillArgumentParser(description="下载并校验远程 HTTPS Excel 文件")
    parser.add_argument(
        "--url",
        "--excel-url",
        "--excel_url",
        dest="url",
        required=True,
        help="远程 HTTPS Excel、CSV 或 TSV 地址",
    )
    parser.add_argument(
        "--output",
        required=True,
        help=(
            "本地输出路径；扩展名必须是 "
            ".xlsx、.xlsm、.xltx、.xltm、.xls、.csv 或 .tsv"
        ),
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=DEFAULT_TIMEOUT_SECONDS,
        help=f"连接和读取超时秒数，默认 {DEFAULT_TIMEOUT_SECONDS}",
    )
    parser.add_argument(
        "--max-bytes",
        "--max_bytes",
        dest="max_bytes",
        type=int,
        default=DEFAULT_MAX_BYTES,
        help=f"最大下载字节数，默认 {DEFAULT_MAX_BYTES}",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="允许覆盖本次任务已存在的缓存文件",
    )
    args = parser.parse_args(argv)

    args.url = _validate_https_url(args.url)
    if args.timeout < 1 or args.timeout > 600:
        raise ValueError("timeout 必须在 1 到 600 秒之间")
    if args.max_bytes < 1 or args.max_bytes > MAX_ALLOWED_BYTES:
        raise ValueError(
            f"max-bytes 必须在 1 到 {MAX_ALLOWED_BYTES} 之间"
        )
    args.output = output_file(
        args.output,
        TABULAR_INPUT_SUFFIXES,
        overwrite=args.overwrite,
    )
    return args


def _safe_archive_name(name: str) -> None:
    candidate = PurePosixPath(name)
    if (
        candidate.is_absolute()
        or not candidate.parts
        or ".." in candidate.parts
        or candidate.parts[0].endswith(":")
    ):
        raise ValueError(f"Excel 压缩包含不安全路径：{name}")


def _zip_member_is_symlink(info: zipfile.ZipInfo) -> bool:
    mode = (info.external_attr >> 16) & 0xFFFF
    return stat.S_ISLNK(mode)


def _parse_content_types(payload: bytes) -> Any:
    try:
        from defusedxml import ElementTree as ET
        from defusedxml.common import DefusedXmlException

        try:
            return ET.fromstring(payload)
        except (ET.ParseError, DefusedXmlException, ValueError) as exc:
            raise ValueError(f"Excel 内容类型 XML 解析失败：{exc}") from exc
    except ModuleNotFoundError:
        from lxml import etree

        parser = etree.XMLParser(
            resolve_entities=False,
            no_network=True,
            recover=False,
            huge_tree=False,
        )
        try:
            return etree.fromstring(payload, parser=parser)
        except (etree.XMLSyntaxError, ValueError) as exc:
            raise ValueError(f"Excel 内容类型 XML 解析失败：{exc}") from exc


def _validate_ooxml_workbook(
    path: Path,
    expected_suffix: str,
) -> dict[str, Any]:
    try:
        with zipfile.ZipFile(path) as archive:
            infos = archive.infolist()
            if len(infos) > MAX_ARCHIVE_MEMBERS:
                raise ValueError("Excel 压缩包成员数量超过安全限制")

            names: set[str] = set()
            total_size = 0
            for info in infos:
                _safe_archive_name(info.filename)
                if _zip_member_is_symlink(info):
                    raise ValueError(
                        f"Excel 压缩包含符号链接：{info.filename}"
                    )
                if info.file_size > MAX_ARCHIVE_MEMBER_BYTES:
                    raise ValueError(
                        f"Excel 压缩包成员过大：{info.filename}"
                    )
                total_size += info.file_size
                if total_size > MAX_ARCHIVE_UNCOMPRESSED_BYTES:
                    raise ValueError("Excel 解压后总大小超过安全限制")
                if info.filename in names:
                    raise ValueError(
                        f"Excel 压缩包含重复成员：{info.filename}"
                    )
                names.add(info.filename)

            required = {"[Content_Types].xml", "xl/workbook.xml"}
            missing = sorted(required - names)
            if missing:
                raise ValueError(
                    "下载内容不是有效的 Excel OOXML 文件；缺少："
                    + "、".join(missing)
                )

            content_types = archive.read("[Content_Types].xml")
    except zipfile.BadZipFile as exc:
        raise ValueError("下载内容不是有效的 Excel OOXML 压缩包") from exc

    root = _parse_content_types(content_types)
    detected_suffix: Optional[str] = None
    for element in root.iter(f"{{{CONTENT_TYPES_NS}}}Override"):
        if element.attrib.get("PartName") == "/xl/workbook.xml":
            detected_suffix = WORKBOOK_CONTENT_TYPES.get(
                element.attrib.get("ContentType", "")
            )
            break
    if detected_suffix is None:
        raise ValueError("下载内容不是受支持的 Excel OOXML 工作簿")
    if detected_suffix != expected_suffix:
        raise ValueError(
            "下载内容的实际格式为 "
            f"{detected_suffix}，但 output 使用了 {expected_suffix}"
        )

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("当前 Python 未加载环境预置的 openpyxl 模块") from exc

    workbook = None
    try:
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=False,
            keep_vba=detected_suffix in {".xlsm", ".xltm"},
            keep_links=False,
        )
        sheet_count = len(workbook.sheetnames)
        if sheet_count < 1:
            raise ValueError("Excel 工作簿不包含工作表")
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError("下载内容不是可解析的 Excel 工作簿") from exc
    finally:
        if workbook is not None:
            workbook.close()

    return {
        "format": detected_suffix.lstrip("."),
        "sheet_count": sheet_count,
        "archive_member_count": len(infos),
        "archive_uncompressed_bytes": total_size,
        "validation": "ooxml-and-openpyxl",
    }


def _validate_legacy_xls(path: Path) -> dict[str, Any]:
    with path.open("rb") as stream:
        magic = stream.read(len(OLE_COMPOUND_MAGIC))
    if magic != OLE_COMPOUND_MAGIC:
        raise ValueError("下载内容不是有效的旧版 Excel 复合文件")
    return {
        "format": "xls",
        "validation": "ole-compound-signature",
    }


def _validate_delimited_text(path: Path, suffix: str) -> dict[str, Any]:
    with path.open("rb") as stream:
        prefix = stream.read(1024 * 1024)
    if not prefix:
        raise ValueError("下载内容为空")

    if prefix.startswith((b"\xff\xfe", b"\xfe\xff")):
        encoding = "utf-16"
        try:
            text = prefix.decode(encoding)
        except UnicodeDecodeError as exc:
            raise ValueError("下载内容不是可解析的 UTF-16 表格文本") from exc
    else:
        if b"\x00" in prefix:
            raise ValueError("下载内容包含二进制空字节，不是文本表格")
        for encoding in ("utf-8-sig", "gb18030"):
            try:
                text = prefix.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError("下载内容不是 UTF-8 或 GB18030 表格文本")

    normalized = text.lstrip("\ufeff\r\n\t ").lower()
    if not normalized:
        raise ValueError("下载的表格文本不包含有效内容")
    if normalized.startswith(("<!doctype html", "<html", "<?xml")):
        raise ValueError("远程服务器返回了网页或 XML，而不是表格文件")

    return {
        "format": suffix.lstrip("."),
        "encoding": encoding,
        "validation": "delimited-text",
    }


def _validate_workbook(path: Path, suffix: str) -> dict[str, Any]:
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _validate_ooxml_workbook(path, suffix)
    if suffix == ".xls":
        return _validate_legacy_xls(path)
    return _validate_delimited_text(path, suffix)


def _download(args: argparse.Namespace) -> dict[str, Any]:
    output: Path = args.output
    request = urllib.request.Request(
        args.url,
        headers={
            "Accept": (
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet,"
                "application/vnd.ms-excel,text/csv,text/tab-separated-values,"
                "application/octet-stream;q=0.9,*/*;q=0.1"
            ),
            "Accept-Encoding": "identity",
            "User-Agent": USER_AGENT,
        },
        method="GET",
    )
    opener = urllib.request.build_opener(HTTPSOnlyRedirectHandler())
    temp_path: Optional[Path] = None

    try:
        with tempfile.NamedTemporaryFile(
            mode="wb",
            prefix=f".{output.stem}.",
            suffix=f".part{output.suffix}",
            dir=str(output.parent),
            delete=False,
        ) as temp_file:
            temp_path = Path(temp_file.name)
            with opener.open(request, timeout=args.timeout) as response:
                _validate_https_url(response.geturl())
                content_length = response.headers.get("Content-Length")
                if content_length:
                    try:
                        expected_bytes = int(content_length)
                    except ValueError:
                        expected_bytes = 0
                    if expected_bytes > args.max_bytes:
                        raise ValueError(
                            "远程文件超过大小限制："
                            f"最多允许 {args.max_bytes} 字节"
                        )

                downloaded_bytes = 0
                while True:
                    chunk = response.read(CHUNK_SIZE)
                    if not chunk:
                        break
                    downloaded_bytes += len(chunk)
                    if downloaded_bytes > args.max_bytes:
                        raise ValueError(
                            "远程文件超过大小限制："
                            f"最多允许 {args.max_bytes} 字节"
                        )
                    temp_file.write(chunk)

            temp_file.flush()
            os.fsync(temp_file.fileno())

        if downloaded_bytes == 0:
            raise ValueError("远程服务器返回了空文件")

        details = _validate_workbook(temp_path, output.suffix.lower())
        publish_file(temp_path, output, overwrite=args.overwrite)
        temp_path = None
        return {
            "path": str(output),
            "size_bytes": downloaded_bytes,
            **details,
        }
    finally:
        if temp_path is not None:
            try:
                temp_path.unlink(missing_ok=True)
            except OSError:
                pass


def _failure_message(exc: Exception) -> str:
    if isinstance(exc, urllib.error.HTTPError):
        return f"下载失败：远程服务器返回 HTTP {exc.code}"
    if isinstance(exc, (TimeoutError, socket.timeout)):
        return "下载失败：连接或读取超时"
    if isinstance(exc, urllib.error.URLError):
        if isinstance(exc.reason, (TimeoutError, socket.timeout)):
            return "下载失败：连接或读取超时"
        return "下载失败：无法访问远程服务器"
    return failure_message(exc)


def main(argv: Optional[list[str]] = None) -> int:
    try:
        args = _parse_args(sys.argv[1:] if argv is None else argv)
        result = _download(args)
    except Exception as exc:
        emit({"ok": False, "error": _failure_message(exc)})
        return 1

    emit({"ok": True, **result})
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
