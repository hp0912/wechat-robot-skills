#!/usr/bin/env python3

from __future__ import annotations

import argparse
import csv
import os
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Optional

from _xlsx_common import (
    TABULAR_INPUT_SUFFIXES,
    SkillArgumentParser,
    input_file,
    openpyxl_load_options,
    output_file,
    publish_file,
    run_cli,
    run_soffice_convert,
)


OUTPUT_SUFFIXES = {".xlsx", ".csv", ".tsv", ".pdf"}
ENCODINGS = {"utf-8", "utf-8-sig", "gb18030", "gbk"}


def _coerce_value(value: str) -> Any:
    stripped = value.strip()
    if not stripped:
        return ""
    lowered = stripped.lower()
    if lowered in {"true", "false"}:
        return lowered == "true"
    try:
        if not (
            (stripped.startswith("0") and len(stripped) > 1 and "." not in stripped)
            or (stripped.startswith("-0") and len(stripped) > 2 and "." not in stripped)
        ):
            return int(stripped)
    except ValueError:
        pass
    try:
        return float(stripped)
    except ValueError:
        return value


def _read_delimited(
    source: Path,
    *,
    encoding: str,
    infer_types: bool,
) -> list[list[Any]]:
    delimiter = "\t" if source.suffix.lower() == ".tsv" else ","
    rows: list[list[Any]] = []
    with source.open("r", encoding=encoding, newline="") as handle:
        for row in csv.reader(handle, delimiter=delimiter):
            rows.append(
                [_coerce_value(value) for value in row] if infer_types else row
            )
    return rows


def _delimited_to_xlsx(
    source: Path,
    destination: Path,
    *,
    sheet_name: str,
    encoding: str,
    infer_types: bool,
) -> int:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    rows = _read_delimited(
        source,
        encoding=encoding,
        infer_types=infer_types,
    )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name[:31] or "Sheet1"
    for row in rows:
        worksheet.append(row)
    if rows:
        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        for cell in worksheet[1]:
            cell.font = Font(name="Arial", bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.row != 1:
                cell.font = Font(name="Arial", size=10)
            if isinstance(cell.value, (datetime, date)):
                cell.number_format = "yyyy-mm-dd"
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.save(destination)
    return len(rows)


def _xlsx_to_delimited(
    source: Path,
    destination: Path,
    *,
    sheet_name: Optional[str],
    formulas: bool,
    encoding: str,
    delimiter: str,
) -> tuple[str, int]:
    from openpyxl import load_workbook

    workbook = load_workbook(
        source,
        data_only=not formulas,
        read_only=True,
        **{
            key: value
            for key, value in openpyxl_load_options(source).items()
            if key != "read_only"
        },
    )
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"工作表不存在：{sheet_name}；可选：{'、'.join(workbook.sheetnames)}"
            )
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.active
    row_count = 0
    with destination.open("w", encoding=encoding, newline="") as handle:
        writer = csv.writer(handle, delimiter=delimiter)
        for row in worksheet.iter_rows(values_only=True):
            writer.writerow(["" if value is None else value for value in row])
            row_count += 1
    workbook.close()
    return worksheet.title, row_count


def _convert_legacy_to_xlsx(
    source: Path,
    directory: Path,
    *,
    timeout: int,
) -> Path:
    converted, _ = run_soffice_convert(
        source,
        target_format="xlsx",
        output_dir=directory,
        timeout=timeout,
        filter_name="Calc MS Excel 2007 XML",
    )
    return converted


def build_parser() -> argparse.ArgumentParser:
    parser = SkillArgumentParser(
        description="在 Excel、CSV、TSV 和 PDF 之间执行受控转换。"
    )
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--sheet", help="导出 CSV/TSV 时选择工作表")
    parser.add_argument(
        "--sheet-name",
        default="Sheet1",
        help="CSV/TSV 转 Excel 时使用的工作表名称",
    )
    parser.add_argument("--encoding", default="utf-8-sig")
    parser.add_argument("--infer-types", action="store_true")
    parser.add_argument(
        "--formulas",
        action="store_true",
        help="导出 CSV/TSV 时输出公式字符串而不是缓存结果",
    )
    parser.add_argument("--timeout", type=int, default=120)
    parser.add_argument("--overwrite", action="store_true")
    return parser


def main() -> dict[str, Any]:
    args = build_parser().parse_args()
    if args.encoding.lower() not in ENCODINGS:
        raise ValueError(f"encoding 仅支持：{'、'.join(sorted(ENCODINGS))}")
    encoding = args.encoding.lower()
    source = input_file(args.input, TABULAR_INPUT_SUFFIXES)
    destination = output_file(
        args.output,
        OUTPUT_SUFFIXES,
        overwrite=args.overwrite,
    )
    if source == destination:
        raise ValueError("输出路径不能与输入文件相同")

    source_suffix = source.suffix.lower()
    destination_suffix = destination.suffix.lower()
    row_count: Optional[int] = None
    selected_sheet: Optional[str] = None

    with tempfile.TemporaryDirectory(prefix="xlsx-convert-") as temp_name:
        temp_dir = Path(temp_name)
        working_source = source
        if source_suffix == ".xls":
            working_source = _convert_legacy_to_xlsx(
                source,
                temp_dir / "legacy",
                timeout=args.timeout,
            )
            source_suffix = ".xlsx"

        descriptor, temp_output_name = tempfile.mkstemp(
            prefix="converted-",
            suffix=destination_suffix,
            dir=str(destination.parent),
        )
        os.close(descriptor)
        temp_output = Path(temp_output_name)
        temp_output.unlink()
        try:
            if destination_suffix == ".xlsx":
                if source_suffix in {".csv", ".tsv"}:
                    row_count = _delimited_to_xlsx(
                        working_source,
                        temp_output,
                        sheet_name=args.sheet_name,
                        encoding=encoding,
                        infer_types=args.infer_types,
                    )
                    selected_sheet = args.sheet_name[:31] or "Sheet1"
                elif source_suffix in {".xltx", ".xltm", ".xlsm"}:
                    converted, _ = run_soffice_convert(
                        working_source,
                        target_format="xlsx",
                        output_dir=temp_dir / "xlsx",
                        timeout=args.timeout,
                        filter_name="Calc MS Excel 2007 XML",
                    )
                    converted.replace(temp_output)
                elif source_suffix == ".xlsx":
                    raise ValueError("输入和输出都是 .xlsx，无需转换")
                else:
                    raise ValueError("该输入格式不能转换为 .xlsx")
            elif destination_suffix in {".csv", ".tsv"}:
                if source_suffix in {".csv", ".tsv"}:
                    raise ValueError(
                        "CSV 与 TSV 互转不属于 Excel 工作簿转换；请先转为 .xlsx"
                    )
                selected_sheet, row_count = _xlsx_to_delimited(
                    working_source,
                    temp_output,
                    sheet_name=args.sheet,
                    formulas=args.formulas,
                    encoding=encoding,
                    delimiter="\t" if destination_suffix == ".tsv" else ",",
                )
            elif destination_suffix == ".pdf":
                if source_suffix in {".csv", ".tsv"}:
                    intermediate = temp_dir / "source.xlsx"
                    _delimited_to_xlsx(
                        working_source,
                        intermediate,
                        sheet_name=args.sheet_name,
                        encoding=encoding,
                        infer_types=args.infer_types,
                    )
                    working_source = intermediate
                converted, _ = run_soffice_convert(
                    working_source,
                    target_format="pdf",
                    output_dir=temp_dir / "pdf",
                    timeout=args.timeout,
                )
                converted.replace(temp_output)
            else:
                raise ValueError("不支持的目标格式")
            publish_file(temp_output, destination, overwrite=args.overwrite)
        finally:
            if temp_output.exists():
                temp_output.unlink()

    return {
        "path": str(destination),
        "source": str(source),
        "source_format": source.suffix.lower(),
        "output_format": destination_suffix,
        "sheet": selected_sheet,
        "row_count": row_count,
        "formulas_exported": bool(args.formulas),
    }


if __name__ == "__main__":
    raise SystemExit(run_cli(main))
