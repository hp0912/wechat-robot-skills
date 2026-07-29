from __future__ import annotations

import argparse
import contextlib
import importlib.util
import io
import json
import os
import sys
import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from unittest import mock


REPOSITORY_ROOT = Path(__file__).resolve().parents[1]
SCRIPT_PATH = (
    REPOSITORY_ROOT
    / "skills"
    / "export-chat-history"
    / "scripts"
    / "export_chat_history.py"
)


def load_script():
    module_name = "_test_export_chat_history"
    spec = importlib.util.spec_from_file_location(module_name, SCRIPT_PATH)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"无法加载测试脚本：{SCRIPT_PATH}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    return module


class FakeCursor:
    def __init__(
        self,
        *,
        fetchone_result=None,
        fetchall_result=None,
    ) -> None:
        self.fetchone_result = fetchone_result
        self.fetchall_result = fetchall_result or []
        self.sql = ""
        self.params = ()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, traceback):
        return False

    def execute(self, sql, params) -> None:
        self.sql = sql
        self.params = params

    def fetchone(self):
        return self.fetchone_result

    def fetchall(self):
        return self.fetchall_result


class FakeConnection:
    def __init__(self, cursors: list[FakeCursor]) -> None:
        self.cursors = list(cursors)
        self.closed = False

    def cursor(self):
        return self.cursors.pop(0)

    def close(self) -> None:
        self.closed = True


class ExportChatHistoryTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.module = load_script()

    def test_resolve_time_range_supports_date_and_today_default(self) -> None:
        now = datetime(
            2026,
            7,
            29,
            15,
            30,
            tzinfo=self.module.SHANGHAI_TZ,
        )
        date_args = argparse.Namespace(
            export_date="2026-07-28",
            start_time="",
            end_time="",
        )
        start, end = self.module._resolve_time_range(date_args, now)
        self.assertEqual(start.isoformat(), "2026-07-28T00:00:00+08:00")
        self.assertEqual(end.isoformat(), "2026-07-29T00:00:00+08:00")

        default_args = argparse.Namespace(
            export_date="",
            start_time="",
            end_time="",
        )
        start, end = self.module._resolve_time_range(default_args, now)
        self.assertEqual(start.isoformat(), "2026-07-29T00:00:00+08:00")
        self.assertEqual(end, now)

    def test_parse_args_rejects_conflicting_or_partial_ranges(self) -> None:
        with self.assertRaisesRegex(ValueError, "不能与"):
            self.module._parse_args(
                [
                    "--date",
                    "2026-07-29",
                    "--start_time",
                    "2026-07-29 09:00",
                    "--end_time",
                    "2026-07-29 10:00",
                ]
            )
        with self.assertRaisesRegex(ValueError, "必须同时提供"):
            self.module._parse_args(
                ["--start_time", "2026-07-29 09:00"]
            )

    def test_query_matches_chat_room_summary_scope(self) -> None:
        rows = [
            {
                "id": 1,
                "sender_wxid": "member",
                "nickname": "成员",
                "message": "你好",
                "created_at": 1_753_776_000,
            }
        ]
        cursor = FakeCursor(fetchall_result=rows)
        connection = FakeConnection([cursor])

        result = self.module._fetch_messages(
            connection,
            "group@chatroom",
            "robot",
            1_753_776_000,
            1_753_862_400,
        )

        self.assertEqual(result, rows)
        self.assertIn("LEFT JOIN chat_room_members", cursor.sql)
        self.assertIn("messages.sender_wxid != %s", cursor.sql)
        self.assertIn("EXTRACTVALUE", cursor.sql)
        self.assertIn("messages.created_at >= %s", cursor.sql)
        self.assertIn("messages.created_at < %s", cursor.sql)
        self.assertEqual(
            cursor.params,
            (
                "group@chatroom",
                "robot",
                1_753_776_000,
                1_753_862_400,
                self.module.MAX_EXPORT_ROWS + 1,
            ),
        )

    def test_excel_text_limit_includes_formula_protection_prefix(self) -> None:
        raw = "=" + ("x" * self.module.EXCEL_CELL_MAX_LENGTH)
        result = self.module._safe_excel_text(raw)

        self.assertEqual(len(result), self.module.EXCEL_CELL_MAX_LENGTH)
        self.assertTrue(result.startswith("'="))
        self.assertTrue(result.endswith("……（内容过长，已截断）"))

    @unittest.skipUnless(
        importlib.util.find_spec("openpyxl"),
        "openpyxl is not installed",
    )
    def test_workbook_contains_metadata_and_escapes_formula_text(self) -> None:
        from openpyxl import load_workbook

        start = datetime(
            2026,
            7,
            29,
            9,
            0,
            tzinfo=self.module.SHANGHAI_TZ,
        )
        end = datetime(
            2026,
            7,
            29,
            10,
            0,
            tzinfo=self.module.SHANGHAI_TZ,
        )
        messages = [
            {
                "id": 1,
                "sender_wxid": "wxid_1",
                "nickname": "@成员",
                "message": "=HYPERLINK(\"https://example.test\")\x00",
                "created_at": int(start.timestamp()),
            }
        ]

        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "history.xlsx"
            self.module._build_workbook(
                output,
                "group@chatroom",
                "测试群",
                start,
                end,
                messages,
                exported_at=end,
            )

            workbook = load_workbook(output, read_only=True, data_only=False)
            self.assertEqual(workbook.sheetnames, ["导出说明", "聊天记录"])
            info_rows = list(workbook["导出说明"].iter_rows(values_only=True))
            self.assertIn(("群聊名称", "测试群"), info_rows)
            history_rows = list(
                workbook["聊天记录"].iter_rows(values_only=True)
            )
            self.assertEqual(
                history_rows[0],
                (
                    "序号",
                    "发送时间",
                    "发送人",
                    "发送人微信 ID",
                    "消息内容",
                ),
            )
            self.assertEqual(history_rows[1][2], "'@成员")
            self.assertEqual(
                history_rows[1][4],
                "'=HYPERLINK(\"https://example.test\")",
            )
            workbook.close()

    def test_main_sends_generated_file_and_removes_it_after_success(self) -> None:
        start = datetime(
            2026,
            7,
            29,
            9,
            0,
            tzinfo=self.module.SHANGHAI_TZ,
        )
        end = datetime(
            2026,
            7,
            29,
            10,
            0,
            tzinfo=self.module.SHANGHAI_TZ,
        )
        messages = [
            {
                "id": 1,
                "sender_wxid": "wxid_1",
                "nickname": "成员",
                "message": "你好",
                "created_at": int(start.timestamp()),
            }
        ]
        connection = FakeConnection([])
        stdout = io.StringIO()

        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "history.xlsx"

            def build_workbook(path, *args, **kwargs):
                del args, kwargs
                path.write_bytes(b"xlsx")

            def send_file(client_port, chat_room_id, path):
                self.assertEqual(client_port, "9100")
                self.assertEqual(chat_room_id, "group@chatroom")
                self.assertTrue(path.is_file())

            patches = (
                mock.patch.object(
                    self.module,
                    "_resolve_time_range",
                    return_value=(start, end),
                ),
                mock.patch.object(
                    self.module,
                    "_ensure_runtime_dependencies",
                ),
                mock.patch.object(
                    self.module,
                    "_mysql_connect",
                    return_value=connection,
                ),
                mock.patch.object(
                    self.module,
                    "_fetch_chat_room_name",
                    return_value="测试群",
                ),
                mock.patch.object(
                    self.module,
                    "_fetch_messages",
                    return_value=messages,
                ),
                mock.patch.object(
                    self.module,
                    "_create_output_path",
                    return_value=output,
                ),
                mock.patch.object(
                    self.module,
                    "_build_workbook",
                    side_effect=build_workbook,
                ),
                mock.patch.object(
                    self.module,
                    "_send_file",
                    side_effect=send_file,
                ),
            )
            environment = {
                "ROBOT_FROM_WX_ID": "group@chatroom",
                "ROBOT_WX_ID": "robot",
                "ROBOT_WECHAT_CLIENT_PORT": "9100",
            }
            with contextlib.ExitStack() as stack:
                for patcher in patches:
                    stack.enter_context(patcher)
                stack.enter_context(
                    mock.patch.dict(os.environ, environment, clear=False)
                )
                stack.enter_context(
                    mock.patch.object(sys, "argv", ["export_chat_history.py"])
                )
                with contextlib.redirect_stdout(stdout):
                    return_code = self.module.main()

            self.assertEqual(return_code, 0)
            self.assertTrue(connection.closed)
            self.assertFalse(output.exists())
            result = json.loads(stdout.getvalue())
            self.assertTrue(result["ok"])
            self.assertEqual(result["message_count"], 1)


if __name__ == "__main__":
    unittest.main()
